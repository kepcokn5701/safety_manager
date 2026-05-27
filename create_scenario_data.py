# -*- coding: utf-8 -*-
"""
경남본부 폭염 상황별 시나리오 가상데이터 - 엑셀 파일 생성
8개 시나리오를 각각 별도 엑셀 파일로 생성
각 파일: 시나리오개요 시트 + 작업현장 시트 + 작업자 시트 + 날씨/알림 시트
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = 'scenario_data'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── 공통 스타일 ──
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 단계별 색상
STAGE_FILLS = {
    '해당없음': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    '관심':     PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid'),
    '주의':     PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid'),
    '경고':     PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid'),
    '위험':     PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid'),
}
STAGE_FONTS = {
    '해당없음': Font(name='맑은 고딕', bold=True, size=10, color='006100'),
    '관심':     Font(name='맑은 고딕', bold=True, size=10, color='000000'),
    '주의':     Font(name='맑은 고딕', bold=True, size=10, color='000000'),
    '경고':     Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF'),
    '위험':     Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF'),
}
VULN_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')


def set_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN_BORDER


def set_cell(ws, row, col, value, font=None, fill=None, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or DATA_FONT
    c.border = THIN_BORDER
    c.alignment = Alignment(vertical='center', wrap_text=wrap)
    if fill:
        c.fill = fill
    return c


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_scenario(filename, info, sites, workers_all, timeline=None):
    """하나의 시나리오 엑셀 파일 생성"""
    wb = openpyxl.Workbook()

    # ── 시트1: 시나리오 개요 ──
    ws1 = wb.active
    ws1.title = '시나리오개요'
    overview_items = [
        ('시나리오 ID', info['id']),
        ('시나리오명', info['name']),
        ('설명', info['desc']),
        ('날짜', info['date']),
        ('기상상황', info['weather_cond']),
        ('폭염단계', info['stage']),
        ('작업현장 수', f"{len(sites)}개소"),
        ('투입 작업자 수', f"{len(workers_all)}명"),
        ('취약작업자 수', f"{sum(1 for w in workers_all if w['is_vulnerable'])}명"),
        ('예상 알림 수', f"{info['expected_alerts']}건"),
        ('핵심 대응사항', info['key_action']),
    ]
    set_col_widths(ws1, [18, 80])
    for r, (label, val) in enumerate(overview_items, 1):
        set_cell(ws1, r, 1, label, font=BOLD_FONT)
        set_cell(ws1, r, 2, val)
        ws1.row_dimensions[r].height = 30 if r < 11 else 45

    # 단계 색상 적용
    stage_name = info.get('stage_key', '해당없음')
    if stage_name in STAGE_FILLS:
        ws1.cell(1, 2).value  # noop
        c = ws1.cell(row=6, column=2)
        c.fill = STAGE_FILLS[stage_name]
        c.font = STAGE_FONTS[stage_name]

    # ── 시트2: 작업현장 ──
    ws2 = wb.create_sheet('작업현장')
    site_headers = [
        'NO', '현장명', '주소', '위도', '경도',
        '작업강도', '옥외여부', '기온(°C)', '습도(%)',
        '풍속(m/s)', '체감온도(°C)', 'WBGT(°C)',
        '폭염단계', '조치사항', 'WBGT 권고사항'
    ]
    set_header(ws2, 1, site_headers)
    set_col_widths(ws2, [5, 30, 35, 10, 10, 10, 8, 8, 8, 8, 12, 10, 10, 50, 50])

    for r, s in enumerate(sites, 2):
        w = s['weather']
        stage = s.get('stage_name', '해당없음')
        sf = STAGE_FILLS.get(stage)
        set_cell(ws2, r, 1, r - 1)
        set_cell(ws2, r, 2, s['name'])
        set_cell(ws2, r, 3, s['address'])
        set_cell(ws2, r, 4, s['lat'])
        set_cell(ws2, r, 5, s['lon'])
        set_cell(ws2, r, 6, s['intensity_kr'])
        set_cell(ws2, r, 7, '옥외' if s['is_outdoor'] else '옥내')
        set_cell(ws2, r, 8, w['temp'])
        set_cell(ws2, r, 9, w['hum'])
        set_cell(ws2, r, 10, w['wind'])
        set_cell(ws2, r, 11, w['app_temp'])
        set_cell(ws2, r, 12, w['wbgt'])
        set_cell(ws2, r, 13, stage, font=STAGE_FONTS.get(stage), fill=sf)
        set_cell(ws2, r, 14, s.get('action', ''))
        set_cell(ws2, r, 15, s.get('wbgt_rec', ''))
        ws2.row_dimensions[r].height = 50

    ws2.auto_filter.ref = f'A1:O{len(sites)+1}'
    ws2.freeze_panes = 'A2'

    # ── 시트3: 작업자 ──
    ws3 = wb.create_sheet('작업자')
    worker_headers = [
        'NO', '성명', '연락처', '소속부서', '작업반',
        '취약여부', '취약사유', '배치현장', '폭염단계',
        '조치사항'
    ]
    set_header(ws3, 1, worker_headers)
    set_col_widths(ws3, [5, 10, 16, 12, 10, 8, 22, 28, 10, 45])

    for r, wr in enumerate(workers_all, 2):
        vfill = VULN_FILL if wr['is_vulnerable'] else None
        stage = wr.get('stage_name', '해당없음')
        sf = STAGE_FILLS.get(stage)
        set_cell(ws3, r, 1, r - 1, fill=vfill)
        set_cell(ws3, r, 2, wr['name'], fill=vfill)
        set_cell(ws3, r, 3, wr['phone'], fill=vfill)
        set_cell(ws3, r, 4, wr['dept'], fill=vfill)
        set_cell(ws3, r, 5, wr['team'], fill=vfill)
        set_cell(ws3, r, 6, 'Y' if wr['is_vulnerable'] else 'N', fill=vfill)
        set_cell(ws3, r, 7, wr.get('vuln_reason', ''), fill=vfill)
        set_cell(ws3, r, 8, wr['site_name'], fill=vfill)
        set_cell(ws3, r, 9, stage, font=STAGE_FONTS.get(stage), fill=sf)
        set_cell(ws3, r, 10, wr.get('action', ''), fill=vfill)
        ws3.row_dimensions[r].height = 28

    ws3.auto_filter.ref = f'A1:J{len(workers_all)+1}'
    ws3.freeze_panes = 'A2'

    # ── 시트4: 타임라인 (있는 경우만) ──
    if timeline:
        ws4 = wb.create_sheet('시간대별변화')
        tl_headers = ['시각', '폭염단계', '기온(°C)', '습도(%)', '풍속(m/s)',
                      '체감온도(°C)', 'WBGT(°C)', '상황설명', '조치사항']
        set_header(ws4, 1, tl_headers)
        set_col_widths(ws4, [8, 12, 8, 8, 8, 12, 10, 45, 50])

        for r, t in enumerate(timeline, 2):
            stage = t.get('stage_name', '해당없음')
            sf = STAGE_FILLS.get(stage)
            set_cell(ws4, r, 1, t['time'])
            set_cell(ws4, r, 2, stage, font=STAGE_FONTS.get(stage), fill=sf)
            set_cell(ws4, r, 3, t['temp'])
            set_cell(ws4, r, 4, t['hum'])
            set_cell(ws4, r, 5, t['wind'])
            set_cell(ws4, r, 6, t['app_temp'])
            set_cell(ws4, r, 7, t['wbgt'])
            set_cell(ws4, r, 8, t['desc'])
            set_cell(ws4, r, 9, t.get('action', ''))
            ws4.row_dimensions[r].height = 35

        ws4.freeze_panes = 'A2'

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f'  생성: {path} ({len(sites)}현장, {len(workers_all)}명)')


# ================================================================
# 시나리오 데이터 정의
# ================================================================

INTENSITY_KR = {'light': '경작업', 'moderate': '중등작업', 'heavy': '중작업', 'very_heavy': '초중작업'}


# ── S01: 평상시 ──
build_scenario(
    'S01_평상시.xlsx',
    info={
        'id': 'S01', 'name': '평상시 (폭염 없음)',
        'desc': '체감온도 30°C 이하의 일반적인 작업일. 전 현장 정상작업 가능. 별도 폭염 대응 불필요.',
        'date': '2026-06-10', 'weather_cond': '맑음, 약한 바람',
        'stage': '해당없음', 'stage_key': '해당없음',
        'expected_alerts': 0,
        'key_action': '정상작업 수행. 기본 안전수칙 준수.'
    },
    sites=[
        {'name': '창원 성산구 가공배전선로 보수공사', 'address': '경상남도 창원시 성산구 중앙대로 210',
         'lat': 35.2284, 'lon': 128.6830, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 27.5, 'hum': 55.0, 'wind': 3.2, 'app_temp': 28.8, 'wbgt': 23.5},
         'stage_name': '해당없음', 'action': '정상작업', 'wbgt_rec': ''},
        {'name': '진주 지중배전선로 신설공사', 'address': '경상남도 진주시 동진로 279',
         'lat': 35.1798, 'lon': 128.1076, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 26.8, 'hum': 58.0, 'wind': 2.8, 'app_temp': 28.2, 'wbgt': 23.1},
         'stage_name': '해당없음', 'action': '정상작업', 'wbgt_rec': ''},
        {'name': '밀양-창녕 345kV 송전선로 점검', 'address': '경상남도 밀양시 가곡7길 30',
         'lat': 35.5036, 'lon': 128.7460, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 26.2, 'hum': 60.0, 'wind': 2.5, 'app_temp': 27.5, 'wbgt': 22.8},
         'stage_name': '해당없음', 'action': '정상작업', 'wbgt_rec': ''},
    ],
    workers_all=[
        {'name': '김진호', 'phone': '010-5701-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 성산구 가공배전선로 보수공사', 'stage_name': '해당없음', 'action': '정상작업'},
        {'name': '박성진', 'phone': '010-5701-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 성산구 가공배전선로 보수공사', 'stage_name': '해당없음', 'action': '정상작업'},
        {'name': '이동규', 'phone': '010-5701-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 성산구 가공배전선로 보수공사', 'stage_name': '해당없음', 'action': '정상작업'},
        {'name': '장세환', 'phone': '010-5704-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '진주 지중배전선로 신설공사', 'stage_name': '해당없음', 'action': '정상작업'},
        {'name': '문태준', 'phone': '010-5704-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '진주 지중배전선로 신설공사', 'stage_name': '해당없음', 'action': '정상작업'},
        {'name': '최상원', 'phone': '010-5704-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '당뇨(만61세)', 'site_name': '진주 지중배전선로 신설공사', 'stage_name': '해당없음', 'action': '정상작업 (건강상태 수시 확인)'},
        {'name': '임창호', 'phone': '010-5712-0001', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '밀양-창녕 345kV 송전선로 점검', 'stage_name': '해당없음', 'action': '정상작업'},
        {'name': '김도현', 'phone': '010-5712-0002', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '밀양-창녕 345kV 송전선로 점검', 'stage_name': '해당없음', 'action': '정상작업'},
        {'name': '최민수', 'phone': '010-5712-0004', 'dept': '송전팀', 'team': '송전2반', 'is_vulnerable': True, 'vuln_reason': '고혈압+당뇨(만62세)', 'site_name': '밀양-창녕 345kV 송전선로 점검', 'stage_name': '해당없음', 'action': '정상작업 (건강상태 수시 확인)'},
    ]
)

# ── S02: 관심 단계 ──
build_scenario(
    'S02_관심단계.xlsx',
    info={
        'id': 'S02', 'name': '관심 단계 (체감온도 33~35°C)',
        'desc': '체감온도 33°C 이상. 작업 시 주의 환기, 충분한 음료수 및 그늘 휴식공간 제공. 작업자 건강상태 수시 확인.',
        'date': '2026-07-15', 'weather_cond': '맑음, 미풍, 습도 높음',
        'stage': '관심', 'stage_key': '관심',
        'expected_alerts': 7,
        'key_action': '주의 환기 알림 발송. 음료수/그늘 휴식공간 확보. 자율 휴식 권장.'
    },
    sites=[
        {'name': '김해 아파트단지 인입선 공사', 'address': '경상남도 김해시 김해대로 2399',
         'lat': 35.2285, 'lon': 128.8892, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 32.0, 'hum': 65.0, 'wind': 1.5, 'app_temp': 34.2, 'wbgt': 27.8},
         'stage_name': '관심', 'action': '주의 환기, 충분한 음료수 및 그늘 휴식공간 제공', 'wbgt_rec': '[중작업] WBGT 27.8°C → 작업 50%/휴식 50% 권고'},
        {'name': '함안 가공배전선로 이설공사', 'address': '경상남도 함안군 함마대로 1490',
         'lat': 35.2730, 'lon': 128.4065, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 31.5, 'hum': 68.0, 'wind': 1.2, 'app_temp': 33.8, 'wbgt': 27.5},
         'stage_name': '관심', 'action': '주의 환기, 자율 휴식 권장', 'wbgt_rec': '[중작업] WBGT 27.5°C → 작업 75%/휴식 25% 권고'},
        {'name': '고성S/S 변전설비 순시점검', 'address': '경상남도 고성군 성내로 162',
         'lat': 34.9737, 'lon': 128.3218, 'intensity_kr': '경작업', 'is_outdoor': False,
         'weather': {'temp': 31.8, 'hum': 66.0, 'wind': 1.8, 'app_temp': 34.0, 'wbgt': 27.2},
         'stage_name': '관심', 'action': '옥내작업 - 영향 제한적, 건강상태 수시 확인', 'wbgt_rec': ''},
    ],
    workers_all=[
        {'name': '채영환', 'phone': '010-5713-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '김해 아파트단지 인입선 공사', 'stage_name': '관심', 'action': '작업 계속, 음료수 섭취 권장'},
        {'name': '신우진', 'phone': '010-5713-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '김해 아파트단지 인입선 공사', 'stage_name': '관심', 'action': '작업 계속'},
        {'name': '정성민', 'phone': '010-5713-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '심혈관질환(만55세)', 'site_name': '김해 아파트단지 인입선 공사', 'stage_name': '관심', 'action': '건강상태 수시 확인, 이상 시 즉시 휴식'},
        {'name': '손민혁', 'phone': '010-5709-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '함안 가공배전선로 이설공사', 'stage_name': '관심', 'action': '작업 계속'},
        {'name': '강병수', 'phone': '010-5709-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '함안 가공배전선로 이설공사', 'stage_name': '관심', 'action': '작업 계속'},
        {'name': '나영진', 'phone': '010-5709-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '신장질환(만56세)', 'site_name': '함안 가공배전선로 이설공사', 'stage_name': '관심', 'action': '건강상태 수시 확인'},
        {'name': '유동현', 'phone': '010-5708-0001', 'dept': '변전운영팀', 'team': '변전1반', 'is_vulnerable': False, 'site_name': '고성S/S 변전설비 순시점검', 'stage_name': '관심', 'action': '옥내작업 - 정상 수행'},
    ]
)

# ── S03: 주의 단계 ──
build_scenario(
    'S03_주의단계_폭염주의보.xlsx',
    info={
        'id': 'S03', 'name': '주의 단계 (폭염주의보 발령)',
        'desc': '체감온도 35~38°C. 폭염주의보 발령. 매시간 10~15분 규칙적 휴식. 취약작업자 옥외작업 전환 배치.',
        'date': '2026-07-22', 'weather_cond': '맑음, 무풍, 고습',
        'stage': '주의', 'stage_key': '주의',
        'expected_alerts': 12,
        'key_action': '매시간 10~15분 휴식 부여. 취약자 작업전환. 냉방시설/그늘 휴식처 확보. 응급조치 체계 점검.'
    },
    sites=[
        {'name': '창원 성산구 배전선로 활선작업', 'address': '경상남도 창원시 성산구 중앙대로 210',
         'lat': 35.2284, 'lon': 128.6830, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 34.0, 'hum': 70.0, 'wind': 0.8, 'app_temp': 36.5, 'wbgt': 30.2},
         'stage_name': '주의', 'action': '매시간 10~15분 휴식, 냉방시설 확보, 응급조치 체계 점검', 'wbgt_rec': '[중작업] WBGT 30.2°C → 작업 25%/휴식 75% 권고'},
        {'name': '거제 옥포 배전설비 보강공사', 'address': '경상남도 거제시 서문로3길 21',
         'lat': 34.8804, 'lon': 128.6214, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 33.5, 'hum': 72.0, 'wind': 1.0, 'app_temp': 35.8, 'wbgt': 29.8},
         'stage_name': '주의', 'action': '매시간 10~15분 휴식', 'wbgt_rec': '[중작업] WBGT 29.8°C → 작업 50%/휴식 50% 권고'},
        {'name': '사천-고성 154kV 송전선로 점검', 'address': '경상남도 사천시 동금6길 28',
         'lat': 35.0028, 'lon': 128.0652, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 34.2, 'hum': 68.0, 'wind': 0.5, 'app_temp': 37.0, 'wbgt': 30.5},
         'stage_name': '주의', 'action': '취약자 작업전환 배치, 매시간 10~15분 휴식', 'wbgt_rec': '[중작업] WBGT 30.5°C → 작업 25%/휴식 75% 권고'},
    ],
    workers_all=[
        {'name': '김진호', 'phone': '010-5701-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 성산구 배전선로 활선작업', 'stage_name': '주의', 'action': '작업 계속 (매시간 10~15분 휴식)'},
        {'name': '박성진', 'phone': '010-5701-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 성산구 배전선로 활선작업', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '이동규', 'phone': '010-5701-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 성산구 배전선로 활선작업', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '하승우', 'phone': '010-5706-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '거제 옥포 배전설비 보강공사', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '구영민', 'phone': '010-5706-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '거제 옥포 배전설비 보강공사', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '남궁현', 'phone': '010-5707-0001', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '사천-고성 154kV 송전선로 점검', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '임성호', 'phone': '010-5707-0002', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '사천-고성 154kV 송전선로 점검', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '정윤석', 'phone': '010-5707-0003', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': True, 'vuln_reason': '고혈압(만58세)', 'site_name': '사천-고성 154kV 송전선로 점검', 'stage_name': '주의', 'action': '★ 즉시 옥외작업 전환 배치'},
        {'name': '김태현', 'phone': '010-5707-0004', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '사천-고성 154kV 송전선로 점검', 'stage_name': '주의', 'action': '작업 계속'},
    ]
)

# ── S04: 경고 단계 ──
build_scenario(
    'S04_경고단계_폭염경보.xlsx',
    info={
        'id': 'S04', 'name': '경고 단계 (폭염경보 발령)',
        'desc': '체감온도 38~41°C. 폭염경보 발령. 14~17시 옥외작업 자제. 취약자 옥외작업 금지. 작업시간 단축.',
        'date': '2026-08-01', 'weather_cond': '쾌청, 무풍, 극고온',
        'stage': '경고', 'stage_key': '경고',
        'expected_alerts': 20,
        'key_action': '14~17시 옥외작업 최대한 자제. 취약자 옥외작업 금지(즉시 철수). 매시간 15~20분 휴식. 비상연락망 재확인.'
    },
    sites=[
        {'name': '밀양-창녕 345kV 송전선로 점검', 'address': '경상남도 밀양시 가곡7길 30',
         'lat': 35.5036, 'lon': 128.7460, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 37.0, 'hum': 65.0, 'wind': 0.3, 'app_temp': 39.8, 'wbgt': 33.2},
         'stage_name': '경고', 'action': '14~17시 옥외작업 자제, 매시간 15~20분 휴식, 취약자 즉시 철수', 'wbgt_rec': '[중작업] WBGT 33.2°C → 작업 즉시 중지 권고'},
        {'name': '하동 배전선로 긴급 보수', 'address': '경상남도 하동군 경서대로 83',
         'lat': 35.0673, 'lon': 127.7516, 'intensity_kr': '초중작업', 'is_outdoor': True,
         'weather': {'temp': 36.5, 'hum': 68.0, 'wind': 0.5, 'app_temp': 39.2, 'wbgt': 32.8},
         'stage_name': '경고', 'action': '긴급작업 - 최소 인원 투입, 14~17시 자제, 매시간 15~20분 휴식', 'wbgt_rec': '[초중작업] WBGT 32.8°C → 작업 즉시 중지 권고'},
        {'name': '남해 해안도로 배전설비 이설', 'address': '경상남도 남해군 남해대로 2962',
         'lat': 34.8378, 'lon': 127.8923, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 35.8, 'hum': 75.0, 'wind': 2.0, 'app_temp': 38.5, 'wbgt': 31.5},
         'stage_name': '경고', 'action': '해풍 있으나 경고단계 유지, 취약자 즉시 철수', 'wbgt_rec': '[중작업] WBGT 31.5°C → 작업 25%/휴식 75% 이상 권고'},
    ],
    workers_all=[
        {'name': '임창호', 'phone': '010-5712-0001', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '밀양-창녕 345kV 송전선로 점검', 'stage_name': '경고', 'action': '14~17시 작업 자제, 매시간 15~20분 휴식'},
        {'name': '김도현', 'phone': '010-5712-0002', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '밀양-창녕 345kV 송전선로 점검', 'stage_name': '경고', 'action': '14~17시 작업 자제'},
        {'name': '최민수', 'phone': '010-5712-0004', 'dept': '송전팀', 'team': '송전2반', 'is_vulnerable': True, 'vuln_reason': '고혈압+당뇨(만62세)', 'site_name': '밀양-창녕 345kV 송전선로 점검', 'stage_name': '경고', 'action': '★★ 옥외작업 금지 → 즉시 철수'},
        {'name': '한상우', 'phone': '010-5712-0005', 'dept': '송전팀', 'team': '송전2반', 'is_vulnerable': False, 'site_name': '밀양-창녕 345kV 송전선로 점검', 'stage_name': '경고', 'action': '14~17시 작업 자제'},
        {'name': '마성규', 'phone': '010-5718-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '하동 배전선로 긴급 보수', 'stage_name': '경고', 'action': '긴급작업 최소인원 투입, 매시간 15~20분 휴식'},
        {'name': '피재호', 'phone': '010-5718-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '하동 배전선로 긴급 보수', 'stage_name': '경고', 'action': '긴급작업 투입'},
        {'name': '감동석', 'phone': '010-5718-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '고혈압(만60세)', 'site_name': '하동 배전선로 긴급 보수', 'stage_name': '경고', 'action': '★★ 옥외작업 금지 → 즉시 철수'},
        {'name': '봉승진', 'phone': '010-5718-0004', 'dept': '배전운영팀', 'team': '배전2반', 'is_vulnerable': False, 'site_name': '하동 배전선로 긴급 보수', 'stage_name': '경고', 'action': '긴급작업 투입'},
        {'name': '복영수', 'phone': '010-5719-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '남해 해안도로 배전설비 이설', 'stage_name': '경고', 'action': '14~17시 작업 자제'},
        {'name': '단기현', 'phone': '010-5719-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '남해 해안도로 배전설비 이설', 'stage_name': '경고', 'action': '14~17시 작업 자제'},
        {'name': '연준석', 'phone': '010-5719-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '심장질환(만57세)', 'site_name': '남해 해안도로 배전설비 이설', 'stage_name': '경고', 'action': '★★ 옥외작업 금지 → 즉시 철수'},
    ]
)

# ── S05: 위험 단계 ──
build_scenario(
    'S05_위험단계_극한폭염.xlsx',
    info={
        'id': 'S05', 'name': '위험 단계 (극한 폭염)',
        'desc': '체감온도 41°C 이상. 옥외 중작업 즉시 중지. 14~17시 옥외작업 원칙적 중지. 필수작업만 최소 인원으로 실시.',
        'date': '2026-08-05', 'weather_cond': '쾌청, 무풍, 역대급 폭염',
        'stage': '위험', 'stage_key': '위험',
        'expected_alerts': 25,
        'key_action': '옥외 중작업 즉시 중지! 14~17시 원칙적 중지. 긴급 정전복구 등 필수작업만 최소인원 교대투입. 매시간 30분 이상 휴식.'
    },
    sites=[
        {'name': '창원 긴급 정전복구 작업', 'address': '경상남도 창원시 성산구 중앙대로 210',
         'lat': 35.2284, 'lon': 128.6830, 'intensity_kr': '초중작업', 'is_outdoor': True,
         'weather': {'temp': 39.5, 'hum': 60.0, 'wind': 0.2, 'app_temp': 42.3, 'wbgt': 35.8},
         'stage_name': '위험', 'action': '★★★ 옥외 중작업 즉시 중지! 긴급복구는 최소인원(2명) 교대투입, 매시간 30분 이상 휴식', 'wbgt_rec': '[초중작업] WBGT 35.8°C → 모든 기준 초과, 작업 즉시 중지'},
        {'name': '진주 배전공사 (작업중지)', 'address': '경상남도 진주시 동진로 279',
         'lat': 35.1798, 'lon': 128.1076, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 39.2, 'hum': 62.0, 'wind': 0.3, 'app_temp': 41.8, 'wbgt': 35.2},
         'stage_name': '위험', 'action': '★★★ 옥외작업 전면 중지 → 전원 대피. 익일 06~10시로 재편성', 'wbgt_rec': '[중작업] WBGT 35.2°C → 작업 즉시 중지'},
        {'name': '마산S/S 변전설비 점검 (옥내)', 'address': '경상남도 창원시 마산합포구 3.15대로 966',
         'lat': 35.1997, 'lon': 128.5736, 'intensity_kr': '중등작업', 'is_outdoor': False,
         'weather': {'temp': 39.0, 'hum': 63.0, 'wind': 0.5, 'app_temp': 41.5, 'wbgt': 34.8},
         'stage_name': '위험', 'action': '옥내작업 유지. 변전소 냉방 가동 확인. 취약자 건강상태 수시 모니터링.', 'wbgt_rec': ''},
    ],
    workers_all=[
        {'name': '김진호', 'phone': '010-5701-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 긴급 정전복구 작업', 'stage_name': '위험', 'action': '긴급복구 최소인원 교대투입, 매시간 30분 이상 휴식'},
        {'name': '박성진', 'phone': '010-5701-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창원 긴급 정전복구 작업', 'stage_name': '위험', 'action': '긴급복구 교대투입'},
        {'name': '장세환', 'phone': '010-5704-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '진주 배전공사 (작업중지)', 'stage_name': '위험', 'action': '★★★ 작업 즉시 중지, 대피'},
        {'name': '문태준', 'phone': '010-5704-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '진주 배전공사 (작업중지)', 'stage_name': '위험', 'action': '★★★ 작업 즉시 중지, 대피'},
        {'name': '최상원', 'phone': '010-5704-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '당뇨(만61세)', 'site_name': '진주 배전공사 (작업중지)', 'stage_name': '위험', 'action': '★★★ 최우선 철수 대상 → 즉시 대피'},
        {'name': '오창석', 'phone': '010-5702-0001', 'dept': '변전정비팀', 'team': '변전1반', 'is_vulnerable': False, 'site_name': '마산S/S 변전설비 점검 (옥내)', 'stage_name': '위험', 'action': '옥내작업 유지, 냉방 확인'},
        {'name': '한재민', 'phone': '010-5702-0002', 'dept': '변전정비팀', 'team': '변전1반', 'is_vulnerable': False, 'site_name': '마산S/S 변전설비 점검 (옥내)', 'stage_name': '위험', 'action': '옥내작업 유지'},
        {'name': '윤기태', 'phone': '010-5702-0003', 'dept': '변전정비팀', 'team': '변전1반', 'is_vulnerable': True, 'vuln_reason': '심장질환(만59세)', 'site_name': '마산S/S 변전설비 점검 (옥내)', 'stage_name': '위험', 'action': '옥내작업 유지, 건강상태 수시 확인'},
    ]
)

# ── S06: 단계 급상승 (타임라인) ──
build_scenario(
    'S06_단계급상승_관심에서경고.xlsx',
    info={
        'id': 'S06', 'name': '단계 급상승 (관심 → 경고)',
        'desc': '오전 관심 단계에서 오후 경고 단계까지 급상승. 시간대별 단계 전환에 따른 실시간 대응 필요. 14:30 경고 도달 시 현장 철수 판단.',
        'date': '2026-07-28', 'weather_cond': '오전 구름 → 오후 쾌청, 습도 급상승',
        'stage': '관심 → 주의 → 경고 → 주의 (하강)', 'stage_key': '경고',
        'expected_alerts': 18,
        'key_action': '시간대별 단계 전환 모니터링. 13:00 주의단계 취약자 전환. 14:30 경고단계 전원 작업자제/취약자 철수. 17:00 하향 시 재개 판단.'
    },
    sites=[
        {'name': '통영 해저케이블 연결 배전공사', 'address': '경상남도 통영시 중림2로 25',
         'lat': 34.8544, 'lon': 128.4336, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 37.0, 'hum': 72.0, 'wind': 0.3, 'app_temp': 40.1, 'wbgt': 33.5},
         'stage_name': '경고', 'action': '14:30 경고단계 도달 → 취약자 즉시 철수, 14~17시 작업 자제', 'wbgt_rec': '[중작업] 피크 시 WBGT 33.5°C → 작업 중지 권고'},
        {'name': '거창-함양 154kV 송전선로 점검', 'address': '경상남도 거창군 강남로 246',
         'lat': 35.6868, 'lon': 127.9099, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 37.0, 'hum': 72.0, 'wind': 0.3, 'app_temp': 40.1, 'wbgt': 33.5},
         'stage_name': '경고', 'action': '산악지역 경고단계 → 취약자 즉시 하산, 전원 14~17시 작업 중지', 'wbgt_rec': '[중작업] 피크 시 WBGT 33.5°C → 작업 중지 권고'},
    ],
    workers_all=[
        {'name': '류건우', 'phone': '010-5705-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '통영 해저케이블 연결 배전공사', 'stage_name': '경고', 'action': '14:30~ 작업 자제, 17:00 하향 시 재개 가능'},
        {'name': '신태영', 'phone': '010-5705-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '통영 해저케이블 연결 배전공사', 'stage_name': '경고', 'action': '14:30~ 작업 자제'},
        {'name': '조현석', 'phone': '010-5705-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '뇌혈관질환(만63세)', 'site_name': '통영 해저케이블 연결 배전공사', 'stage_name': '경고', 'action': '★ 13:00 주의단계 시 작업전환, 14:30 경고 시 즉시 철수'},
        {'name': '변성호', 'phone': '010-5715-0001', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '거창-함양 154kV 송전선로 점검', 'stage_name': '경고', 'action': '14:30~ 전원 작업 중지, 하산'},
        {'name': '곽동현', 'phone': '010-5715-0002', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '거창-함양 154kV 송전선로 점검', 'stage_name': '경고', 'action': '14:30~ 작업 중지'},
        {'name': '성재민', 'phone': '010-5715-0003', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': True, 'vuln_reason': '뇌혈관질환(만64세)', 'site_name': '거창-함양 154kV 송전선로 점검', 'stage_name': '경고', 'action': '★ 13:00 주의단계 시 즉시 하산 조치'},
        {'name': '우영석', 'phone': '010-5715-0004', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '거창-함양 154kV 송전선로 점검', 'stage_name': '경고', 'action': '14:30~ 작업 중지'},
    ],
    timeline=[
        {'time': '09:00', 'stage_name': '해당없음', 'temp': 29.0, 'hum': 55.0, 'wind': 2.5, 'app_temp': 30.5, 'wbgt': 24.8, 'desc': '작업 시작, 정상 기온', 'action': '정상작업 개시'},
        {'time': '11:00', 'stage_name': '관심', 'temp': 32.0, 'hum': 62.0, 'wind': 1.5, 'app_temp': 33.8, 'wbgt': 27.5, 'desc': '기온 상승, 체감온도 33°C 돌파', 'action': '주의 환기 알림 발송, 음료수 섭취 권장'},
        {'time': '13:00', 'stage_name': '주의', 'temp': 35.0, 'hum': 68.0, 'wind': 0.8, 'app_temp': 37.2, 'wbgt': 30.8, 'desc': '체감온도 35°C 돌파, 폭염주의보', 'action': '★ 취약자 작업전환 알림! 매시간 10~15분 휴식 시행'},
        {'time': '14:30', 'stage_name': '경고', 'temp': 37.0, 'hum': 72.0, 'wind': 0.3, 'app_temp': 40.1, 'wbgt': 33.5, 'desc': '체감온도 38°C 돌파, 폭염경보 상향', 'action': '★★ 14~17시 옥외작업 자제! 취약자 옥외작업 금지, 현장 철수 지시'},
        {'time': '17:00', 'stage_name': '주의', 'temp': 34.5, 'hum': 65.0, 'wind': 1.2, 'app_temp': 36.8, 'wbgt': 30.2, 'desc': '기온 하강, 경고→주의 하향', 'action': '작업 재개 가능 알림 (매시간 10~15분 휴식 유지)'},
        {'time': '18:00', 'stage_name': '관심', 'temp': 32.5, 'hum': 60.0, 'wind': 1.8, 'app_temp': 34.2, 'wbgt': 27.8, 'desc': '작업 종료 시간대, 관심 단계로 하향', 'action': '작업 종료, 잔여 정리작업 가능'},
    ]
)

# ── S07: 취약작업자 집중관리 ──
build_scenario(
    'S07_취약작업자_집중관리.xlsx',
    info={
        'id': 'S07', 'name': '취약작업자 집중관리',
        'desc': '주의 단계(35~36°C)에서 취약작업자가 다수 배치된 현장. 취약자 전원 작업전환 시 인력 부족 문제 발생. 대체인력 배치 판단 필요.',
        'date': '2026-07-20', 'weather_cond': '맑음, 미풍',
        'stage': '주의 (취약자 강화기준 적용)', 'stage_key': '주의',
        'expected_alerts': 15,
        'key_action': '취약자 전원 옥외작업 전환. 인력부족 현장(2인1조 미달) 대체인력 긴급 배치. 작업일정 재조정 검토.'
    },
    sites=[
        {'name': '밀양 송전선로 점검 (취약자 1명)', 'address': '경상남도 밀양시 가곡7길 30',
         'lat': 35.5036, 'lon': 128.7460, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 34.0, 'hum': 68.0, 'wind': 1.0, 'app_temp': 36.2, 'wbgt': 30.0},
         'stage_name': '주의', 'action': '취약자 1명 전환 → 실작업 4명, 작업 계속 가능', 'wbgt_rec': '[중작업] WBGT 30.0°C → 작업 25%/휴식 75% 권고'},
        {'name': '김해 인입선 공사 (취약자 1명)', 'address': '경상남도 김해시 김해대로 2399',
         'lat': 35.2285, 'lon': 128.8892, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 33.8, 'hum': 70.0, 'wind': 0.8, 'app_temp': 35.9, 'wbgt': 29.6},
         'stage_name': '주의', 'action': '★ 취약자 전환 시 2명 → 최소인원 미달! 추가인력 요청 필요', 'wbgt_rec': '[중작업] WBGT 29.6°C → 작업 50%/휴식 50% 권고'},
        {'name': '창녕 배전선로 점검 (취약자 1명)', 'address': '경상남도 창녕군 남창녕로 71',
         'lat': 35.5445, 'lon': 128.4924, 'intensity_kr': '중등작업', 'is_outdoor': True,
         'weather': {'temp': 33.5, 'hum': 66.0, 'wind': 1.2, 'app_temp': 35.5, 'wbgt': 29.2},
         'stage_name': '주의', 'action': '★★ 2인1조 중 취약자 전환 → 단독작업 불가! 대체인력 배치 필수', 'wbgt_rec': '[중등작업] WBGT 29.2°C → 작업 75%/휴식 25% 권고'},
    ],
    workers_all=[
        {'name': '임창호', 'phone': '010-5712-0001', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '밀양 송전선로 점검 (취약자 1명)', 'stage_name': '주의', 'action': '작업 계속 (매시간 10~15분 휴식)'},
        {'name': '김도현', 'phone': '010-5712-0002', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '밀양 송전선로 점검 (취약자 1명)', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '박정훈', 'phone': '010-5712-0003', 'dept': '송전팀', 'team': '송전1반', 'is_vulnerable': False, 'site_name': '밀양 송전선로 점검 (취약자 1명)', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '최민수', 'phone': '010-5712-0004', 'dept': '송전팀', 'team': '송전2반', 'is_vulnerable': True, 'vuln_reason': '고혈압+당뇨(만62세)', 'site_name': '밀양 송전선로 점검 (취약자 1명)', 'stage_name': '주의', 'action': '★ 즉시 옥외작업 전환 → 차량 대기'},
        {'name': '한상우', 'phone': '010-5712-0005', 'dept': '송전팀', 'team': '송전2반', 'is_vulnerable': False, 'site_name': '밀양 송전선로 점검 (취약자 1명)', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '채영환', 'phone': '010-5713-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '김해 인입선 공사 (취약자 1명)', 'stage_name': '주의', 'action': '작업 계속 → 대체인력 도착까지 안전작업만 수행'},
        {'name': '신우진', 'phone': '010-5713-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '김해 인입선 공사 (취약자 1명)', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '정성민', 'phone': '010-5713-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '심혈관질환(만55세)', 'site_name': '김해 인입선 공사 (취약자 1명)', 'stage_name': '주의', 'action': '★ 즉시 옥외작업 전환 → 차량 대기'},
        {'name': '강도윤', 'phone': '010-5711-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '창녕 배전선로 점검 (취약자 1명)', 'stage_name': '주의', 'action': '★★ 단독작업 불가 → 대체인력 도착까지 작업 중지'},
        {'name': '홍석준', 'phone': '010-5711-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '호흡기질환(만60세)', 'site_name': '창녕 배전선로 점검 (취약자 1명)', 'stage_name': '주의', 'action': '★ 즉시 옥외작업 전환'},
    ]
)

# ── S08: 지역별 기온 편차 ──
build_scenario(
    'S08_지역별_기온편차.xlsx',
    info={
        'id': 'S08', 'name': '지역별 기온 편차 (내륙 vs 해안)',
        'desc': '같은 날 경남 내륙(합천·창녕)은 경고 단계, 해안(통영·거제)은 주의 단계. 지역별 차등 대응 필요. 해풍 유입으로 해안 2~3°C 낮음.',
        'date': '2026-08-03', 'weather_cond': '내륙: 쾌청 무풍 / 해안: 맑음 약한 해풍',
        'stage': '내륙 경고 / 해안 주의', 'stage_key': '경고',
        'expected_alerts': 22,
        'key_action': '내륙(합천·창녕): 경고 대응(14~17시 작업자제, 취약자 금지). 해안(통영·거제): 주의 대응(매시간 10~15분 휴식, 취약자 전환).'
    },
    sites=[
        {'name': '[내륙] 합천 광케이블 포설공사', 'address': '경상남도 합천군 대야로 921',
         'lat': 35.5660, 'lon': 128.1657, 'intensity_kr': '중등작업', 'is_outdoor': True,
         'weather': {'temp': 37.5, 'hum': 58.0, 'wind': 0.3, 'app_temp': 39.5, 'wbgt': 32.8},
         'stage_name': '경고', 'action': '내륙 분지지형 + 무풍 → 경고. 14~17시 작업 자제, 맨홀 내부 온도 별도 측정', 'wbgt_rec': '[중등작업] WBGT 32.8°C → 작업 중지 권고'},
        {'name': '[내륙] 창녕 배전선로 점검', 'address': '경상남도 창녕군 남창녕로 71',
         'lat': 35.5445, 'lon': 128.4924, 'intensity_kr': '중등작업', 'is_outdoor': True,
         'weather': {'temp': 37.8, 'hum': 60.0, 'wind': 0.2, 'app_temp': 40.2, 'wbgt': 33.5},
         'stage_name': '경고', 'action': '★ 취약자 즉시 철수, 단독작업 불가 → 대체인력 또는 작업 연기', 'wbgt_rec': '[중등작업] WBGT 33.5°C → 작업 중지 권고'},
        {'name': '[해안] 통영 배전선로 공사', 'address': '경상남도 통영시 중림2로 25',
         'lat': 34.8544, 'lon': 128.4336, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 33.5, 'hum': 72.0, 'wind': 3.5, 'app_temp': 35.8, 'wbgt': 29.0},
         'stage_name': '주의', 'action': '해풍 유입으로 주의 단계 유지. 취약자 작업전환, 매시간 10~15분 휴식', 'wbgt_rec': '[중작업] WBGT 29.0°C → 작업 50%/휴식 50% 권고'},
        {'name': '[해안] 거제 배전설비 보강', 'address': '경상남도 거제시 서문로3길 21',
         'lat': 34.8804, 'lon': 128.6214, 'intensity_kr': '중작업', 'is_outdoor': True,
         'weather': {'temp': 33.2, 'hum': 74.0, 'wind': 3.8, 'app_temp': 35.2, 'wbgt': 28.5},
         'stage_name': '주의', 'action': '해풍 유입으로 주의 단계. 일반 근로자만 배치 → 매시간 10~15분 휴식', 'wbgt_rec': '[중작업] WBGT 28.5°C → 작업 50%/휴식 50% 권고'},
    ],
    workers_all=[
        {'name': '탁재욱', 'phone': '010-5714-0001', 'dept': 'ICT팀', 'team': 'ICT1반', 'is_vulnerable': False, 'site_name': '[내륙] 합천 광케이블 포설공사', 'stage_name': '경고', 'action': '14~17시 작업 자제, 매시간 15~20분 휴식'},
        {'name': '남현수', 'phone': '010-5714-0002', 'dept': 'ICT팀', 'team': 'ICT1반', 'is_vulnerable': False, 'site_name': '[내륙] 합천 광케이블 포설공사', 'stage_name': '경고', 'action': '14~17시 작업 자제'},
        {'name': '정호진', 'phone': '010-5714-0003', 'dept': 'ICT팀', 'team': 'ICT1반', 'is_vulnerable': False, 'site_name': '[내륙] 합천 광케이블 포설공사', 'stage_name': '경고', 'action': '14~17시 작업 자제'},
        {'name': '강도윤', 'phone': '010-5711-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '[내륙] 창녕 배전선로 점검', 'stage_name': '경고', 'action': '★ 대체인력 배치까지 작업 중지'},
        {'name': '홍석준', 'phone': '010-5711-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '호흡기질환(만60세)', 'site_name': '[내륙] 창녕 배전선로 점검', 'stage_name': '경고', 'action': '★★ 옥외작업 금지 → 즉시 철수'},
        {'name': '류건우', 'phone': '010-5705-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '[해안] 통영 배전선로 공사', 'stage_name': '주의', 'action': '작업 계속 (매시간 10~15분 휴식)'},
        {'name': '신태영', 'phone': '010-5705-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '[해안] 통영 배전선로 공사', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '조현석', 'phone': '010-5705-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': True, 'vuln_reason': '뇌혈관질환(만63세)', 'site_name': '[해안] 통영 배전선로 공사', 'stage_name': '주의', 'action': '★ 즉시 옥외작업 전환'},
        {'name': '하승우', 'phone': '010-5706-0001', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '[해안] 거제 배전설비 보강', 'stage_name': '주의', 'action': '작업 계속 (매시간 10~15분 휴식)'},
        {'name': '구영민', 'phone': '010-5706-0002', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '[해안] 거제 배전설비 보강', 'stage_name': '주의', 'action': '작업 계속'},
        {'name': '안재혁', 'phone': '010-5706-0003', 'dept': '배전운영팀', 'team': '배전1반', 'is_vulnerable': False, 'site_name': '[해안] 거제 배전설비 보강', 'stage_name': '주의', 'action': '작업 계속'},
    ]
)

print(f'\n총 8개 시나리오 엑셀 파일 생성 완료! ({OUTPUT_DIR}/ 디렉토리)')
